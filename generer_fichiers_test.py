# generer_fichiers_test.py
import openpyxl

# Articles — fichier valide avec 1 erreur + 1 avertissement volontaires
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['CPN', 'Désignation', 'Famille achats', 'Catégorie', 'Client', 'Site', 'Obsolète'])
ws.append(['00540A711A', 'Relais électromécanique', 'Électromécanique', 'Relais', 'SAFRAN', 'TRONICO', 'Non'])
ws.append(['00999TEST1', 'Article test import', '', 'Test', 'THALES', 'TRONICO', 'Non'])  # famille vide -> avertissement
ws.append(['', 'Sans CPN', 'Passif', 'Test', 'SAFRAN', 'TRONICO', 'Non'])  # CPN manquant -> erreur
wb.save('test_articles.xlsx')

# Articles — colonnes manquantes (doit renvoyer 400)
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.append(['CPN', 'Nom du produit'])  # mauvaises colonnes
ws2.append(['00999TEST2', 'Test colonnes invalides'])
wb2.save('test_articles_mauvaises_colonnes.xlsx')

# Fournisseurs — fichier valide avec 1 erreur + 1 avertissement
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.append(['Nom', 'Contact', 'Actif'])
ws3.append(['Fournisseur Test SA', 'contact@test.com', 'Oui'])
ws3.append(['Fournisseur Sans Contact', '', 'Oui'])  # contact vide -> avertissement
ws3.append(['', 'orphelin@test.com', 'Oui'])  # nom manquant -> erreur
wb3.save('test_fournisseurs.xlsx')

print("3 fichiers générés dans le dossier courant.")